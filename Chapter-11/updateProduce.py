#! python3
# updateProduce.py - corrects costs in the produce sales spreadsheet
# you need the produceSales.xlsx file to run this program

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

#the produce types and their updated prices:

PRICE_UPDATES = {'Garlic': 3.07,
                 'Celery': 1.19,
                 'Lemon': 1.27}

#look through the rows and update prices

for rowNum in range(2, sheet.get_highest_row()): #skip first row of col names
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

wb.save('updatedProduceSales.xlsx')

print('Prices are now updated!')
